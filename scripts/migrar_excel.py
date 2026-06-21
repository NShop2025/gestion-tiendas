"""Migra el histórico de NeptunoShop desde el Excel original a Postgres/Supabase.

Uso:
    python scripts/migrar_excel.py "C:/ruta/a/Libro.xlsx" --tienda neptunoshop [--dry-run]

Requiere DATABASE_URL en el entorno (.env) y que el schema (db/migrations) ya esté aplicado.

Cubre: Ventas NeptunoShop, Ventas WEB, Ventas Tiendimport, Compras (+ gastos en
columnas N:P), Retiros, FLEX MP, FLEX SANTANDER (tabla principal de envíos).

NO cubre casos sueltos como la mini-tabla "COMPRAS" de FLEX SANTANDER (columnas T:V,
una sola fila histórica) — esos se cargan a mano desde la app una vez migrado el resto.
"""

import argparse
import datetime
import os
import re
import sys
from datetime import date

import openpyxl
from dotenv import load_dotenv
from openpyxl.utils.datetime import to_excel
from sqlalchemy import create_engine, text


def _fecha(valor):
    """Convierte celda de fecha a date(). Si es un string tipeado mal (ej '0t/09/2025'),
    intenta rescatar mes/año por regex con día aproximado (15), para no perder el registro."""
    if hasattr(valor, "date"):
        return valor.date()
    if isinstance(valor, str):
        m = re.search(r"(\d{1,2})/(\d{4})$", valor)
        if m:
            mes, anio = int(m.group(1)), int(m.group(2))
            return date(anio, mes, 15)
    return None


def _como_numero(valor):
    """Convierte un string tipo '$325' o '1.234,5' a float. Devuelve None si no se puede."""
    texto = str(valor).strip().replace("$", "").replace(" ", "")
    if not texto:
        return None
    try:
        return float(texto)
    except ValueError:
        return None


def _numero(valor):
    """Corrige celdas numéricas que Excel guardó como fecha/hora por un formato de celda
    equivocado (ej: alguien tipeó un precio en una celda con formato de hora), o como texto
    con formato de moneda (ej: '$325')."""
    if isinstance(valor, (datetime.datetime, datetime.time)):
        return to_excel(valor)
    if isinstance(valor, str):
        return _como_numero(valor) or 0
    return valor or 0


def _es_numero_valido(valor) -> bool:
    """False para placeholders tipo '-' o texto que no representan una cantidad real."""
    if isinstance(valor, (int, float)):
        return True
    if isinstance(valor, (datetime.datetime, datetime.time)):
        return True
    if isinstance(valor, str):
        return _como_numero(valor) is not None
    return False


def conectar_tienda(conn, slug: str) -> str:
    fila = conn.execute(text("select id from tiendas where slug = :slug"), {"slug": slug}).fetchone()
    if not fila:
        raise SystemExit(f"No existe la tienda con slug '{slug}'. Corré primero db/migrations/002_seed_tiendas.sql")
    return str(fila[0])


def obtener_o_crear_producto(conn, tienda_id: str, nombre: str, cache: dict) -> str:
    nombre = (nombre or "").strip()
    if not nombre:
        return None
    if nombre in cache:
        return cache[nombre]
    fila = conn.execute(
        text("select id from productos where tienda_id = :t and nombre = :n"),
        {"t": tienda_id, "n": nombre},
    ).fetchone()
    if fila:
        cache[nombre] = str(fila[0])
        return cache[nombre]
    nueva = conn.execute(
        text("insert into productos (tienda_id, nombre) values (:t, :n) returning id"),
        {"t": tienda_id, "n": nombre},
    ).fetchone()
    cache[nombre] = str(nueva[0])
    return cache[nombre]


def migrar_ventas(conn, wb, hoja: str, canal: str, tienda_id: str, productos_cache: dict, dry_run: bool) -> int:
    ws = wb[hoja]
    filas = []
    for r in range(2, ws.max_row + 1):
        fecha = ws.cell(r, 1).value
        producto_nombre = ws.cell(r, 3).value
        cantidad = ws.cell(r, 4).value
        # cantidad <= 0 son devoluciones/cancelaciones cargadas como venta de 0 unidades:
        # no aportan ingreso ni costo, se descartan en vez de migrarlas como venta inválida.
        if fecha is None or not producto_nombre or not _es_numero_valido(cantidad) or _numero(cantidad) <= 0:
            continue

        hora = ws.cell(r, 2).value
        hora = hora if hasattr(hora, "hour") else None

        precio_unitario = _numero(ws.cell(r, 5).value)
        comision_ml = _numero(ws.cell(r, 7).value)
        ingreso_envio = _numero(ws.cell(r, 9).value)
        costo_total = _numero(ws.cell(r, 10).value)
        comentario = ws.cell(r, 12).value

        costo_unitario_venta = float(costo_total) / float(cantidad) if cantidad else 0

        producto_id = obtener_o_crear_producto(conn, tienda_id, str(producto_nombre), productos_cache)

        filas.append(
            {
                "tienda_id": tienda_id,
                "producto_id": producto_id,
                "canal": canal,
                "fecha": fecha.date() if hasattr(fecha, "date") else fecha,
                "hora": hora,
                "cantidad": cantidad,
                "precio_unitario": precio_unitario,
                "comision_ml": comision_ml,
                "ingreso_envio": ingreso_envio,
                "costo_unitario_venta": costo_unitario_venta,
                "comentario": str(comentario) if comentario else None,
            }
        )

    if not dry_run and filas:
        conn.execute(
            text(
                """
                insert into ventas
                    (tienda_id, producto_id, canal, fecha, hora, cantidad, precio_unitario,
                     comision_ml, ingreso_envio, costo_unitario_venta, comentario)
                values
                    (:tienda_id, :producto_id, :canal, :fecha, :hora, :cantidad, :precio_unitario,
                     :comision_ml, :ingreso_envio, :costo_unitario_venta, :comentario)
                """
            ),
            filas,
        )
    return len(filas)


def migrar_compras(conn, wb, tienda_id: str, productos_cache: dict, dry_run: bool) -> int:
    ws = wb["Compras"]
    filas = []
    for r in range(2, ws.max_row + 1):
        fecha = _fecha(ws.cell(r, 1).value)
        producto_nombre = ws.cell(r, 2).value
        cantidad = ws.cell(r, 3).value
        costo_unitario = ws.cell(r, 4).value
        # cantidad negativa = devolucion a proveedor (resta del costo total), cantidad 0 = fila
        # incompleta o fila de encabezado/total mezclada en los datos: se descarta. Filas sin
        # fecha valida ('-') con costo 0 tambien se descartan (no aportan plata).
        if fecha is None or not producto_nombre or not _es_numero_valido(cantidad) or _numero(cantidad) == 0:
            continue

        producto_id = obtener_o_crear_producto(conn, tienda_id, str(producto_nombre), productos_cache)
        filas.append(
            {
                "tienda_id": tienda_id,
                "producto_id": producto_id,
                "fecha": fecha,
                "cantidad": _numero(cantidad),
                "costo_unitario": _numero(costo_unitario),
                "proveedor_comentario": str(ws.cell(r, 6).value) if ws.cell(r, 6).value else None,
            }
        )

    if not dry_run and filas:
        conn.execute(
            text(
                """
                insert into compras (tienda_id, producto_id, fecha, cantidad, costo_unitario, proveedor_comentario)
                values (:tienda_id, :producto_id, :fecha, :cantidad, :costo_unitario, :proveedor_comentario)
                """
            ),
            filas,
        )
    return len(filas)


def migrar_gastos_compras(conn, wb, tienda_id: str, dry_run: bool) -> int:
    """Mini-tabla de gastos dentro de la hoja Compras, columnas N (fecha), O (concepto), P (monto)."""
    ws = wb["Compras"]
    filas = []
    for r in range(2, ws.max_row + 1):
        fecha = _fecha(ws.cell(r, 14).value)
        concepto = ws.cell(r, 15).value
        monto = ws.cell(r, 16).value
        if fecha is None or not concepto or not _es_numero_valido(monto):
            continue
        filas.append(
            {
                "tienda_id": tienda_id,
                "fecha": fecha,
                "concepto": str(concepto),
                "monto": _numero(monto),
                "cuenta": "mercado_pago",
                "comentario": str(ws.cell(r, 17).value) if ws.cell(r, 17).value else None,
            }
        )

    if not dry_run and filas:
        conn.execute(
            text(
                """
                insert into gastos (tienda_id, fecha, concepto, monto, cuenta, comentario)
                values (:tienda_id, :fecha, :concepto, :monto, :cuenta, :comentario)
                """
            ),
            filas,
        )
    return len(filas)


def migrar_retiros(conn, wb, tienda_id: str, dry_run: bool) -> int:
    ws = wb["Retiros"]
    filas = []
    for r in range(2, ws.max_row + 1):
        fecha = ws.cell(r, 1).value
        monto = ws.cell(r, 2).value
        socio = ws.cell(r, 3).value
        # monto negativo = pago/reembolso recibido del socio (resta del total retirado).
        if fecha is None or not _es_numero_valido(monto) or _numero(monto) == 0 or not socio:
            continue
        filas.append(
            {
                "tienda_id": tienda_id,
                "fecha": fecha.date() if hasattr(fecha, "date") else fecha,
                "monto": _numero(monto),
                "socio": str(socio),
                "comentario": str(ws.cell(r, 4).value) if ws.cell(r, 4).value else None,
            }
        )

    if not dry_run and filas:
        conn.execute(
            text(
                "insert into retiros (tienda_id, fecha, monto, socio, comentario) "
                "values (:tienda_id, :fecha, :monto, :socio, :comentario)"
            ),
            filas,
        )
    return len(filas)


def migrar_envios(conn, wb, hoja: str, cuenta: str, col_fecha: int, tienda_id: str, dry_run: bool) -> int:
    ws = wb[hoja]
    filas = []
    for r in range(2, ws.max_row + 1):
        fecha = ws.cell(r, col_fecha).value
        cadete = ws.cell(r, col_fecha + 1).value
        cantidad = ws.cell(r, col_fecha + 2).value
        costo_unitario = ws.cell(r, col_fecha + 3).value
        if fecha is None or not cadete or not _es_numero_valido(cantidad):
            continue
        filas.append(
            {
                "tienda_id": tienda_id,
                "cuenta": cuenta,
                "fecha": fecha.date() if hasattr(fecha, "date") else fecha,
                "cadete": str(cadete),
                "cantidad_envios": _numero(cantidad),
                "costo_unitario": _numero(costo_unitario),
                "comentario": None,
            }
        )

    if not dry_run and filas:
        conn.execute(
            text(
                """
                insert into envios (tienda_id, cuenta, fecha, cadete, cantidad_envios, costo_unitario, comentario)
                values (:tienda_id, :cuenta, :fecha, :cadete, :cantidad_envios, :costo_unitario, :comentario)
                """
            ),
            filas,
        )
    return len(filas)


def ajustar_stock_negativo(conn, tienda_id: str, dry_run: bool) -> int:
    """Cierra a 0 el stock de productos que quedaron negativos por el cruce de nombres
    entre Compras y Ventas en el Excel original (revisado a mano con Facu el 2026-06-21).

    Inserta una compra de ajuste con costo_unitario = 0, así no toca ningún total de dinero
    (compras, ganancia, saldo), solo corrige la cantidad para que el stock cierre en 0.
    """
    negativos = conn.execute(
        text(
            "select producto_id, producto, stock_actual from stock_actual "
            "where tienda_id = :tienda_id and stock_actual < 0"
        ),
        {"tienda_id": tienda_id},
    ).fetchall()

    if not dry_run:
        for producto_id, producto, stock_actual in negativos:
            conn.execute(
                text(
                    """
                    insert into compras (tienda_id, producto_id, fecha, cantidad, costo_unitario, proveedor_comentario)
                    values (:tienda_id, :producto_id, :fecha, :cantidad, 0, :comentario)
                    """
                ),
                {
                    "tienda_id": tienda_id,
                    "producto_id": producto_id,
                    "fecha": date.today(),
                    "cantidad": -stock_actual,
                    "comentario": "Ajuste stock a 0 - probable nombre distinto entre Compras y Ventas en el Excel original",
                },
            )

    return len(negativos)


def main():
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("excel_path")
    parser.add_argument("--tienda", default="neptunoshop", help="slug de la tienda destino")
    parser.add_argument("--dry-run", action="store_true", help="solo cuenta filas, no inserta nada")
    args = parser.parse_args()

    load_dotenv()
    database_url = os.environ.get("DATABASE_URL")
    if not database_url:
        sys.exit("Falta DATABASE_URL en el entorno (.env)")

    engine = create_engine(database_url)
    wb = openpyxl.load_workbook(args.excel_path, data_only=True)
    productos_cache: dict = {}

    with engine.begin() as conn:
        tienda_id = conectar_tienda(conn, args.tienda)

        n_neptuno = migrar_ventas(conn, wb, "Ventas NeptunoShop", "mercado_libre", tienda_id, productos_cache, args.dry_run)
        n_web = migrar_ventas(conn, wb, "Ventas WEB", "web", tienda_id, productos_cache, args.dry_run)
        n_tiendimport = migrar_ventas(conn, wb, "Ventas Tiendimport", "otro", tienda_id, productos_cache, args.dry_run)
        n_compras = migrar_compras(conn, wb, tienda_id, productos_cache, args.dry_run)
        n_gastos = migrar_gastos_compras(conn, wb, tienda_id, args.dry_run)
        n_retiros = migrar_retiros(conn, wb, tienda_id, args.dry_run)
        n_flex_mp = migrar_envios(conn, wb, "FLEX MP", "mercado_pago", 8, tienda_id, args.dry_run)
        n_flex_santander = migrar_envios(conn, wb, "FLEX SANTANDER", "santander", 7, tienda_id, args.dry_run)
        n_ajustes = ajustar_stock_negativo(conn, tienda_id, args.dry_run)

        if args.dry_run:
            conn.rollback()

    print(f"Ventas Mercado Libre: {n_neptuno}")
    print(f"Ventas Web:           {n_web}")
    print(f"Ventas Tiendimport:   {n_tiendimport}")
    print(f"Compras:              {n_compras}")
    print(f"Gastos (Compras N:P): {n_gastos}")
    print(f"Retiros:              {n_retiros}")
    print(f"Envíos FLEX MP:       {n_flex_mp}")
    print(f"Envíos FLEX Santander:{n_flex_santander}")
    print(f"Ajustes de stock a 0: {n_ajustes}")
    if args.dry_run:
        print("\n(dry-run: no se insertó nada)")


if __name__ == "__main__":
    main()

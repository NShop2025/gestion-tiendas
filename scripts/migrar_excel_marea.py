"""Migra el histórico de Marea Boutique desde su Excel a Postgres/Supabase.

Uso:
    python scripts/migrar_excel_marea.py "C:/ruta/a/Libro - Marea.xlsx" [--dry-run]

Requiere DATABASE_URL en el entorno (.env) y que el schema (db/migrations) ya esté aplicado,
con la tienda 'marea' sembrada (002_seed_tiendas.sql).

A diferencia del Excel de NeptunoShop, el de Marea tiene otra estructura de hojas/columnas
(hoja "Ventas " en vez de "Ventas NeptunoShop", sin "Ventas WEB"/"Ventas Tiendimport" con datos,
sin "FLEX SANTANDER") y solo tiene datos reales cargados en "Compras" y "Ventas " (ambas del
20/6/2026). Las hojas FLEX MP / FLEX traen una tabla pivot con fechas de 2025 que no corresponde
a Marea (empezó a vender en 2026) y se descarta deliberadamente. Retiros, Ventas WEB y
Ventas Tiendimport están vacías en este Excel.

Cubre: Compras, Ventas (canal web).
"""

import argparse
import os
import sys

import openpyxl
from dotenv import load_dotenv
from sqlalchemy import create_engine, text


def _numero(valor):
    if valor is None:
        return 0
    return valor


def conectar_tienda(conn, slug: str) -> str:
    fila = conn.execute(text("select id from tiendas where slug = :slug"), {"slug": slug}).fetchone()
    if not fila:
        raise SystemExit(f"No existe la tienda con slug '{slug}'. Corré primero db/migrations/002_seed_tiendas.sql")
    return str(fila[0])


def obtener_o_crear_producto(conn, tienda_id: str, nombre: str, cache: dict) -> str:
    nombre = (nombre or "").strip()
    if not nombre:
        return None
    if nombre in cache:
        return cache[nombre]
    fila = conn.execute(
        text("select id from productos where tienda_id = :t and nombre = :n"),
        {"t": tienda_id, "n": nombre},
    ).fetchone()
    if fila:
        cache[nombre] = str(fila[0])
        return cache[nombre]
    nueva = conn.execute(
        text("insert into productos (tienda_id, nombre) values (:t, :n) returning id"),
        {"t": tienda_id, "n": nombre},
    ).fetchone()
    cache[nombre] = str(nueva[0])
    return cache[nombre]


def migrar_compras(conn, wb, tienda_id: str, productos_cache: dict, dry_run: bool) -> int:
    """Hoja 'Compras': A=Fecha, B=Producto, C=Cantidad, D=Costo Unitario.
    La columna E (Costo Total) está vacía en todo el Excel de Marea, así que
    costo_total se calcula como cantidad x costo_unitario (a diferencia del Excel
    de NeptunoShop, donde costo_total era un valor escrito a mano).

    cuenta='otra': Marea no tuvo capital inicial propio - esta mercadería ya se pagó
    y se descontó como compra en la caja de NeptunoShop. Estas filas solo sirven para
    registrar cantidad/costo de cada producto (stock y margen de venta), no deben volver
    a restar plata del saldo de Marea (saldo_disponible excluye cuenta='otra')."""
    ws = wb["Compras"]
    filas = []
    for r in range(2, ws.max_row + 1):
        fecha = ws.cell(r, 1).value
        producto_nombre = ws.cell(r, 2).value
        cantidad = ws.cell(r, 3).value
        costo_unitario = ws.cell(r, 4).value

        if fecha is None or not producto_nombre or not isinstance(cantidad, (int, float)) or cantidad == 0:
            continue

        costo_unitario = _numero(costo_unitario)
        costo_total = cantidad * costo_unitario

        producto_id = obtener_o_crear_producto(conn, tienda_id, str(producto_nombre), productos_cache)
        filas.append(
            {
                "tienda_id": tienda_id,
                "producto_id": producto_id,
                "fecha": fecha.date() if hasattr(fecha, "date") else fecha,
                "cantidad": cantidad,
                "costo_unitario": costo_unitario,
                "costo_total": costo_total,
                "cuenta": "otra",
                "proveedor_comentario": str(ws.cell(r, 6).value) if ws.cell(r, 6).value else None,
            }
        )

    if not dry_run and filas:
        conn.execute(
            text(
                """
                insert into compras
                    (tienda_id, producto_id, fecha, cantidad, costo_unitario, costo_total, cuenta, proveedor_comentario)
                values
                    (:tienda_id, :producto_id, :fecha, :cantidad, :costo_unitario, :costo_total, :cuenta, :proveedor_comentario)
                """
            ),
            filas,
        )
    return len(filas)


def migrar_ventas(conn, wb, tienda_id: str, productos_cache: dict, dry_run: bool) -> int:
    """Hoja 'Ventas ': A=Fecha, B=Hora, C=Producto, D=Cantidad, E=Precio Unitario,
    H=Ingreso Neto, J=Costo Total. La columna I (Ingreso Envío por fila) está vacía
    en todo el Excel de Marea, a diferencia de NeptunoShop donde sí tenía datos."""
    ws = wb["Ventas "]
    filas = []
    for r in range(2, ws.max_row + 1):
        fecha = ws.cell(r, 1).value
        producto_nombre = ws.cell(r, 3).value
        cantidad = ws.cell(r, 4).value
        if fecha is None or not producto_nombre or not isinstance(cantidad, (int, float)) or cantidad <= 0:
            continue

        hora = ws.cell(r, 2).value
        hora = hora if hasattr(hora, "hour") else None

        precio_unitario = _numero(ws.cell(r, 5).value)
        costo_total = _numero(ws.cell(r, 10).value)
        comentario = ws.cell(r, 12).value

        ingreso_bruto = float(precio_unitario) * float(cantidad)
        ingreso_neto_real = _numero(ws.cell(r, 8).value)
        comision_ml = ingreso_bruto - ingreso_neto_real

        costo_unitario_venta = float(costo_total) / float(cantidad) if cantidad else 0

        producto_id = obtener_o_crear_producto(conn, tienda_id, str(producto_nombre), productos_cache)
        filas.append(
            {
                "tienda_id": tienda_id,
                "producto_id": producto_id,
                "canal": "web",
                "fecha": fecha.date() if hasattr(fecha, "date") else fecha,
                "hora": hora,
                "cantidad": cantidad,
                "precio_unitario": precio_unitario,
                "comision_ml": comision_ml,
                "ingreso_envio": 0,
                "costo_unitario_venta": costo_unitario_venta,
                "comentario": str(comentario) if comentario else None,
            }
        )

    if not dry_run and filas:
        conn.execute(
            text(
                """
                insert into ventas
                    (tienda_id, producto_id, canal, fecha, hora, cantidad, precio_unitario,
                     comision_ml, ingreso_envio, costo_unitario_venta, comentario)
                values
                    (:tienda_id, :producto_id, :canal, :fecha, :hora, :cantidad, :precio_unitario,
                     :comision_ml, :ingreso_envio, :costo_unitario_venta, :comentario)
                """
            ),
            filas,
        )
    return len(filas)


def main():
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("excel_path")
    parser.add_argument("--dry-run", action="store_true", help="solo cuenta filas, no inserta nada")
    args = parser.parse_args()

    load_dotenv()
    database_url = os.environ.get("DATABASE_URL")
    if not database_url:
        sys.exit("Falta DATABASE_URL en el entorno (.env)")

    engine = create_engine(database_url)
    wb = openpyxl.load_workbook(args.excel_path, data_only=True)
    productos_cache: dict = {}

    with engine.begin() as conn:
        tienda_id = conectar_tienda(conn, "marea")
        n_compras = migrar_compras(conn, wb, tienda_id, productos_cache, args.dry_run)
        n_ventas = migrar_ventas(conn, wb, tienda_id, productos_cache, args.dry_run)

        if args.dry_run:
            conn.rollback()

    print(f"Compras: {n_compras}")
    print(f"Ventas:  {n_ventas}")
    if args.dry_run:
        print("\n(dry-run: no se insertó nada)")


if __name__ == "__main__":
    main()

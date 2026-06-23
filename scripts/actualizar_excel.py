"""Carga incremental: agrega a la base solo las ventas, compras y gastos del Excel de
NeptunoShop que todavía NO están en la base (por fecha+detalle), para no duplicar lo que
ya se migró antes o se cargó a mano desde la app.

Uso:
    python scripts/actualizar_excel.py "C:/ruta/al/Libro.xlsx" --dry-run
    python scripts/actualizar_excel.py "C:/ruta/al/Libro.xlsx"
"""

import argparse
import os
import sys

import openpyxl
from dotenv import load_dotenv
from sqlalchemy import create_engine, text

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from scripts.categorizar_gastos import MAPEO as MAPEO_CATEGORIAS
from scripts.migrar_excel import (
    _es_numero_valido,
    _fecha,
    _numero,
    conectar_tienda,
    obtener_o_crear_producto,
)


def _hora_key(hora):
    return hora.strftime("%H:%M:%S") if hora is not None else None


def candidatos_ventas(wb, hoja: str, canal: str) -> list[dict]:
    ws = wb[hoja]
    filas = []
    for r in range(2, ws.max_row + 1):
        fecha = ws.cell(r, 1).value
        producto_nombre = ws.cell(r, 3).value
        cantidad = ws.cell(r, 4).value
        if fecha is None or not producto_nombre or not _es_numero_valido(cantidad) or _numero(cantidad) <= 0:
            continue

        hora = ws.cell(r, 2).value
        hora = hora if hasattr(hora, "hour") else None

        precio_unitario = _numero(ws.cell(r, 5).value)
        ingreso_envio = _numero(ws.cell(r, 9).value)
        costo_total = _numero(ws.cell(r, 10).value)
        comentario = ws.cell(r, 12).value

        ingreso_bruto = float(precio_unitario) * float(cantidad)
        valor_h = ws.cell(r, 8).value
        ingreso_neto_real = _numero(valor_h) if _es_numero_valido(valor_h) else 0
        comision_ml = ingreso_bruto - ingreso_neto_real
        costo_unitario_venta = float(costo_total) / float(cantidad) if cantidad else 0

        filas.append(
            {
                "producto_nombre": str(producto_nombre).strip(),
                "canal": canal,
                "fecha": fecha.date() if hasattr(fecha, "date") else fecha,
                "hora": hora,
                "cantidad": cantidad,
                "precio_unitario": precio_unitario,
                "comision_ml": comision_ml,
                "ingreso_envio": ingreso_envio,
                "costo_unitario_venta": costo_unitario_venta,
                "comentario": str(comentario) if comentario else None,
            }
        )
    return filas


def candidatos_compras(wb) -> list[dict]:
    ws = wb["Compras"]
    filas = []
    for r in range(2, ws.max_row + 1):
        fecha = _fecha(ws.cell(r, 1).value)
        producto_nombre = ws.cell(r, 2).value
        cantidad = ws.cell(r, 3).value
        costo_unitario = ws.cell(r, 4).value
        costo_total_escrito = ws.cell(r, 5).value

        if fecha is None:
            continue

        tiene_costo = _es_numero_valido(costo_total_escrito) and _numero(costo_total_escrito) != 0
        tiene_producto = bool(producto_nombre) and not (
            isinstance(producto_nombre, str) and producto_nombre.strip().upper().startswith("TOTAL")
        )
        if not tiene_producto and not tiene_costo:
            continue

        if tiene_producto:
            nombre = str(producto_nombre).strip()
        else:
            proveedor = ws.cell(r, 6).value
            nombre = f"Compra sin detalle ({proveedor})" if proveedor else "Compra sin detalle"

        cantidad_valida = _es_numero_valido(cantidad) and _numero(cantidad) != 0
        cant = _numero(cantidad) if cantidad_valida else 1
        costo_total = _numero(costo_total_escrito) if tiene_costo else 0
        costo_unit = _numero(costo_unitario) if _es_numero_valido(costo_unitario) else 0

        filas.append(
            {
                "producto_nombre": nombre,
                "fecha": fecha,
                "cantidad": cant,
                "costo_unitario": costo_unit,
                "costo_total": costo_total,
                "cuenta": "mercado_pago",
                "proveedor_comentario": str(ws.cell(r, 6).value) if ws.cell(r, 6).value else None,
            }
        )
    return filas


def candidatos_gastos(wb) -> list[dict]:
    ws = wb["Compras"]
    filas = []
    for r in range(2, ws.max_row + 1):
        fecha = _fecha(ws.cell(r, 14).value)
        concepto = ws.cell(r, 15).value
        monto = ws.cell(r, 16).value
        if fecha is None or not concepto or not _es_numero_valido(monto):
            continue
        filas.append(
            {
                "fecha": fecha,
                "concepto": str(concepto),
                "monto": _numero(monto),
                "cuenta": "mercado_pago",
                "categoria": MAPEO_CATEGORIAS.get(str(concepto), "otros"),
                "comentario": str(ws.cell(r, 17).value) if ws.cell(r, 17).value else None,
            }
        )
    return filas


def candidatos_envios(wb, hoja: str, cuenta: str, col_fecha: int) -> list[dict]:
    ws = wb[hoja]
    filas = []
    for r in range(2, ws.max_row + 1):
        fecha = ws.cell(r, col_fecha).value
        cadete = ws.cell(r, col_fecha + 1).value
        cantidad = ws.cell(r, col_fecha + 2).value
        costo_unitario = ws.cell(r, col_fecha + 3).value
        if fecha is None or not cadete or not _es_numero_valido(cantidad):
            continue
        filas.append(
            {
                "cuenta": cuenta,
                "fecha": fecha.date() if hasattr(fecha, "date") else fecha,
                "cadete": str(cadete).strip(),
                "cantidad_envios": _numero(cantidad),
                "costo_unitario": _numero(costo_unitario),
                "comentario": None,
            }
        )
    return filas


def candidatos_gastos_flex_santander(wb) -> list[dict]:
    """Mini-tabla 'COMPRAS' dentro de FLEX SANTANDER (columnas T:V), gastos pagados con
    Santander que no pasan por la mini-tabla de Compras!N:Q."""
    ws = wb["FLEX SANTANDER"]
    filas = []
    for r in range(3, ws.max_row + 1):
        fecha = _fecha(ws.cell(r, 20).value)
        monto = ws.cell(r, 21).value
        proveedor = ws.cell(r, 22).value
        if fecha is None or not _es_numero_valido(monto):
            continue
        concepto_txt = str(proveedor).strip() if proveedor else "(sin concepto)"
        filas.append(
            {
                "fecha": fecha,
                "concepto": concepto_txt,
                "monto": _numero(monto),
                "cuenta": "santander",
                "categoria": MAPEO_CATEGORIAS.get(concepto_txt, "otros"),
                "comentario": None,
            }
        )
    return filas


def main():
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("excel_path")
    parser.add_argument("--tienda", default="neptunoshop")
    parser.add_argument("--dry-run", action="store_true")
    args = parser.parse_args()

    load_dotenv()
    database_url = os.environ.get("DATABASE_URL")
    if not database_url:
        sys.exit("Falta DATABASE_URL en el entorno (.env)")

    engine = create_engine(database_url)
    wb = openpyxl.load_workbook(args.excel_path, data_only=True)
    productos_cache: dict = {}

    with engine.begin() as conn:
        tienda_id = conectar_tienda(conn, args.tienda)

        # --- Ventas ---
        existentes_ventas = {
            (row.fecha, _hora_key(row.hora), row.nombre.strip(), float(row.cantidad), float(row.precio_unitario), row.canal)
            for row in conn.execute(
                text(
                    """
                    select v.fecha, v.hora, p.nombre, v.cantidad, v.precio_unitario, v.canal
                    from ventas v join productos p on p.id = v.producto_id
                    where v.tienda_id = :t
                    """
                ),
                {"t": tienda_id},
            )
        }
        candidatas_ventas = (
            candidatos_ventas(wb, "Ventas NeptunoShop", "mercado_libre")
            + candidatos_ventas(wb, "Ventas WEB", "web")
            + candidatos_ventas(wb, "Ventas Tiendimport", "otro")
        )
        nuevas_ventas = [
            f
            for f in candidatas_ventas
            if (
                f["fecha"],
                _hora_key(f["hora"]),
                f["producto_nombre"],
                float(f["cantidad"]),
                float(f["precio_unitario"]),
                f["canal"],
            )
            not in existentes_ventas
        ]

        # La tabla ventas no permite precio_unitario negativo (no hay "ventas negativas" reales).
        # El Excel a veces usa una fila "DIFERENCIAS" en negativo para anular otro ajuste positivo
        # y que la cuenta quede en $0 - si la excluimos, excluimos también su contraparte exacta
        # para no alterar ese neteo.
        negativas = [f for f in nuevas_ventas if f["precio_unitario"] < 0]
        if negativas:
            print(f"Excluidas {len(negativas)} venta(s) con precio negativo (no permitido):")
            for f in negativas:
                print(f"  {f['fecha']} {f['producto_nombre']!r} ${f['precio_unitario']} ({f['comentario']})")
            excluir = set(id(f) for f in negativas)
            for neg in negativas:
                contraparte = [
                    f
                    for f in nuevas_ventas
                    if id(f) not in excluir
                    and f["producto_nombre"] == neg["producto_nombre"]
                    and float(f["precio_unitario"]) == -float(neg["precio_unitario"])
                ]
                for f in contraparte:
                    print(f"  (y su contraparte: {f['fecha']} ${f['precio_unitario']} ({f['comentario']}))")
                    excluir.add(id(f))
            nuevas_ventas = [f for f in nuevas_ventas if id(f) not in excluir]
            print()

        # --- Compras ---
        existentes_compras = {
            (row.fecha, row.nombre.strip(), float(row.cantidad), float(row.costo_unitario), float(row.costo_total))
            for row in conn.execute(
                text(
                    """
                    select c.fecha, p.nombre, c.cantidad, c.costo_unitario, c.costo_total
                    from compras c join productos p on p.id = c.producto_id
                    where c.tienda_id = :t
                    """
                ),
                {"t": tienda_id},
            )
        }
        candidatas_compras = candidatos_compras(wb)
        nuevas_compras = [
            f
            for f in candidatas_compras
            if (
                f["fecha"],
                f["producto_nombre"],
                float(f["cantidad"]),
                float(f["costo_unitario"]),
                float(f["costo_total"]),
            )
            not in existentes_compras
        ]

        # --- Gastos ---
        existentes_gastos = {
            (row.fecha, row.concepto, float(row.monto))
            for row in conn.execute(
                text("select fecha, concepto, monto from gastos where tienda_id = :t"),
                {"t": tienda_id},
            )
        }
        candidatos_gastos_ = candidatos_gastos(wb) + candidatos_gastos_flex_santander(wb)
        nuevos_gastos = [
            f for f in candidatos_gastos_ if (f["fecha"], f["concepto"], float(f["monto"])) not in existentes_gastos
        ]

        # --- Envíos ---
        existentes_envios = {
            (row.fecha, row.cuenta, row.cadete.strip(), float(row.cantidad_envios), float(row.costo_unitario))
            for row in conn.execute(
                text("select fecha, cuenta, cadete, cantidad_envios, costo_unitario from envios where tienda_id = :t"),
                {"t": tienda_id},
            )
        }
        candidatos_envios_ = candidatos_envios(wb, "FLEX MP", "mercado_pago", 8) + candidatos_envios(
            wb, "FLEX SANTANDER", "santander", 7
        )
        nuevos_envios = [
            f
            for f in candidatos_envios_
            if (f["fecha"], f["cuenta"], f["cadete"], float(f["cantidad_envios"]), float(f["costo_unitario"]))
            not in existentes_envios
        ]

        print(f"Ventas:  {len(candidatas_ventas)} en el Excel, {len(nuevas_ventas)} nuevas")
        print(f"Compras: {len(candidatas_compras)} en el Excel, {len(nuevas_compras)} nuevas")
        print(f"Gastos:  {len(candidatos_gastos_)} en el Excel, {len(nuevos_gastos)} nuevos")
        print(f"Envíos:  {len(candidatos_envios_)} en el Excel, {len(nuevos_envios)} nuevos")
        print()

        if nuevas_ventas:
            print("--- Ventas nuevas (primeras 20) ---")
            for f in nuevas_ventas[:20]:
                print(f"  {f['fecha']} {f['hora']}  {f['producto_nombre']!r}  x{f['cantidad']}  ${f['precio_unitario']}  ({f['canal']})")
            if len(nuevas_ventas) > 20:
                print(f"  ... y {len(nuevas_ventas) - 20} más")
            print()

        if nuevas_compras:
            print("--- Compras nuevas (primeras 20) ---")
            for f in nuevas_compras[:20]:
                print(f"  {f['fecha']}  {f['producto_nombre']!r}  x{f['cantidad']}  costo unit ${f['costo_unitario']}  total ${f['costo_total']}")
            if len(nuevas_compras) > 20:
                print(f"  ... y {len(nuevas_compras) - 20} más")
            print()

        if nuevos_gastos:
            print("--- Gastos nuevos ---")
            for f in nuevos_gastos:
                print(f"  {f['fecha']}  {f['concepto']!r}  ${f['monto']}  cuenta={f['cuenta']}  -> categoria: {f['categoria']}")
            print()

        if nuevos_envios:
            print("--- Envíos nuevos ---")
            for f in nuevos_envios:
                print(f"  {f['fecha']}  {f['cadete']!r}  x{f['cantidad_envios']}  ${f['costo_unitario']}/u  cuenta={f['cuenta']}")
            print()

        if args.dry_run:
            print("(dry-run: no se insertó nada)")
            conn.rollback()
            return

        for f in nuevas_ventas:
            producto_id = obtener_o_crear_producto(conn, tienda_id, f["producto_nombre"], productos_cache)
            conn.execute(
                text(
                    """
                    insert into ventas
                        (tienda_id, producto_id, canal, fecha, hora, cantidad, precio_unitario,
                         comision_ml, ingreso_envio, costo_unitario_venta, comentario)
                    values
                        (:tienda_id, :producto_id, :canal, :fecha, :hora, :cantidad, :precio_unitario,
                         :comision_ml, :ingreso_envio, :costo_unitario_venta, :comentario)
                    """
                ),
                {**f, "tienda_id": tienda_id, "producto_id": producto_id},
            )

        for f in nuevas_compras:
            producto_id = obtener_o_crear_producto(conn, tienda_id, f["producto_nombre"], productos_cache)
            conn.execute(
                text(
                    """
                    insert into compras
                        (tienda_id, producto_id, fecha, cantidad, costo_unitario, costo_total, cuenta, proveedor_comentario)
                    values
                        (:tienda_id, :producto_id, :fecha, :cantidad, :costo_unitario, :costo_total, :cuenta, :proveedor_comentario)
                    """
                ),
                {**f, "tienda_id": tienda_id, "producto_id": producto_id},
            )

        for f in nuevos_gastos:
            conn.execute(
                text(
                    """
                    insert into gastos (tienda_id, fecha, concepto, monto, cuenta, categoria, comentario)
                    values (:tienda_id, :fecha, :concepto, :monto, :cuenta, :categoria, :comentario)
                    """
                ),
                {**f, "tienda_id": tienda_id},
            )

        for f in nuevos_envios:
            conn.execute(
                text(
                    """
                    insert into envios (tienda_id, cuenta, fecha, cadete, cantidad_envios, costo_unitario, comentario)
                    values (:tienda_id, :cuenta, :fecha, :cadete, :cantidad_envios, :costo_unitario, :comentario)
                    """
                ),
                {**f, "tienda_id": tienda_id},
            )

    print(
        f"Insertado: {len(nuevas_ventas)} ventas, {len(nuevas_compras)} compras, "
        f"{len(nuevos_gastos)} gastos, {len(nuevos_envios)} envíos."
    )


if __name__ == "__main__":
    main()

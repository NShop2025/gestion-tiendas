"""Consolida la cuenta Santander de NeptunoShop dentro de la caja única.

Una sola vez: carga los gastos/compras que se pagaron con Santander (columna U de la hoja
FLEX SANTANDER del Excel, que la migración inicial no cubría) y elimina la transferencia
interna Mercado Pago -> Santander (no es ingreso ni egreso real en una caja consolidada).

Uso:
    python scripts/consolidar_santander.py "C:/ruta/Libro.xlsx" [--dry-run]
"""

import argparse
import os
import sys

import openpyxl
from dotenv import load_dotenv
from sqlalchemy import create_engine, text


def main():
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("excel_path")
    parser.add_argument("--dry-run", action="store_true")
    args = parser.parse_args()

    load_dotenv()
    database_url = os.environ.get("DATABASE_URL")
    if not database_url:
        sys.exit("Falta DATABASE_URL")

    wb = openpyxl.load_workbook(args.excel_path, data_only=True)
    ws = wb["FLEX SANTANDER"]

    # Columna T (20) = fecha, U (21) = monto, V (22) = concepto/proveedor
    gastos = []
    for r in range(3, 1011):
        fecha = ws.cell(r, 20).value
        monto = ws.cell(r, 21).value
        concepto = ws.cell(r, 22).value
        if not isinstance(monto, (int, float)) or monto == 0:
            continue
        gastos.append(
            {
                "fecha": fecha.date() if hasattr(fecha, "date") else fecha,
                "monto": monto,
                "concepto": str(concepto) if concepto else "Gasto Santander",
            }
        )

    engine = create_engine(database_url)
    with engine.begin() as conn:
        tienda_id = conn.execute(
            text("select id from tiendas where slug = 'neptunoshop'")
        ).fetchone()[0]

        # Evitar duplicar si se corre dos veces.
        ya = conn.execute(
            text(
                "select count(*) from gastos where tienda_id = :t and cuenta = 'santander' "
                "and concepto <> 'envio'"
            ),
            {"t": tienda_id},
        ).scalar()

        if not args.dry_run:
            if ya == 0:
                for g in gastos:
                    conn.execute(
                        text(
                            "insert into gastos (tienda_id, fecha, concepto, monto, cuenta, comentario) "
                            "values (:t, :f, :c, :m, 'santander', 'Pagado con Santander (consolidado)')"
                        ),
                        {"t": tienda_id, "f": g["fecha"], "c": g["concepto"], "m": g["monto"]},
                    )

            # Eliminar la transferencia interna MP->Santander (las dos puntas):
            conn.execute(
                text(
                    "delete from gastos where tienda_id = :t "
                    "and concepto = 'Transferencia a Santander (saldo inicial)'"
                ),
                {"t": tienda_id},
            )
            conn.execute(
                text(
                    "delete from ventas where tienda_id = :t and canal = 'otro' "
                    "and comentario ilike '%SALDO INICIAL SANTANDER%'"
                ),
                {"t": tienda_id},
            )

    print(f"Gastos Santander a cargar: {len(gastos)} (suma {sum(g['monto'] for g in gastos):,.0f})")
    print(f"Ya existían cargados: {ya}")
    if args.dry_run:
        print("\n(dry-run: no se modificó nada)")


if __name__ == "__main__":
    main()
